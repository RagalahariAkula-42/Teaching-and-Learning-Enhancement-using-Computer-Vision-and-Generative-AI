from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.http import HttpResponse
from django.utils.timezone import now, make_aware
from .models import *
from django.db import transaction
from datetime import datetime
from .quiz import extract_text_from_file, Question_mcqs_generator
import openpyxl
from openpyxl.styles import Alignment, Font
import json
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO
import subprocess
from django.contrib import messages
import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from googletrans import Translator
import json
from .summary import video_summary_pipeline
import tempfile

# Path to your service account key file
SERVICE_ACCOUNT_FILE = './quizzz/gen-lang-client-0733632548-8c86d1997c2e.json'
SCOPES = ['https://www.googleapis.com/auth/drive']
# Authenticate using the service account
credentials = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES)
service = build('drive', 'v3', credentials=credentials)


def summary(request):
    summary_text = None  # Initialize summary text variable
    
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']

        # Create a temporary file to handle audio/video processing
        with tempfile.NamedTemporaryFile(delete=True, suffix=uploaded_file.name.split('.')[-1]) as temp_file:
            for chunk in uploaded_file.chunks():
                temp_file.write(chunk)
            temp_file.flush()  # Ensure data is written to the file
            
            # Process the file to get summary
            summary_text = video_summary_pipeline(temp_file.name)

    return render(request, 'summary.html', {'summary_text': summary_text})


# LANGUAGES dictionary to pass to the template
LANGUAGES = {
    'af': 'afrikaans', 'sq': 'albanian', 'am': 'amharic', 'ar': 'arabic', 'hy': 'armenian', 'az': 'azerbaijani',
    'eu': 'basque', 'be': 'belarusian', 'bn': 'bengali', 'bs': 'bosnian', 'bg': 'bulgarian', 'ca': 'catalan',
    'ceb': 'cebuano', 'ny': 'chichewa', 'zh-cn': 'chinese (simplified)', 'zh-tw': 'chinese (traditional)', 'co': 'corsican',
    'hr': 'croatian', 'cs': 'czech', 'da': 'danish', 'nl': 'dutch', 'en': 'english', 'eo': 'esperanto', 'et': 'estonian',
    'tl': 'filipino', 'fi': 'finnish', 'fr': 'french', 'fy': 'frisian', 'gl': 'galician', 'ka': 'georgian', 'de': 'german',
    'el': 'greek', 'gu': 'gujarati', 'ht': 'haitian creole', 'ha': 'hausa', 'haw': 'hawaiian', 'iw': 'hebrew', 'he': 'hebrew',
    'hi': 'hindi', 'hmn': 'hmong', 'hu': 'hungarian', 'is': 'icelandic', 'ig': 'igbo', 'id': 'indonesian', 'ga': 'irish',
    'it': 'italian', 'ja': 'japanese', 'jw': 'javanese', 'kn': 'kannada', 'kk': 'kazakh', 'km': 'khmer', 'ko': 'korean',
    'ku': 'kurdish (kurmanji)', 'ky': 'kyrgyz', 'lo': 'lao', 'la': 'latin', 'lv': 'latvian', 'lt': 'lithuanian',
    'lb': 'luxembourgish', 'mk': 'macedonian', 'mg': 'malagasy', 'ms': 'malay', 'ml': 'malayalam', 'mt': 'maltese',
    'mi': 'maori', 'mr': 'marathi', 'mn': 'mongolian', 'my': 'myanmar (burmese)', 'ne': 'nepali', 'no': 'norwegian',
    'or': 'odia', 'ps': 'pashto', 'fa': 'persian', 'pl': 'polish', 'pt': 'portuguese', 'pa': 'punjabi', 'ro': 'romanian',
    'ru': 'russian', 'sm': 'samoan', 'gd': 'scots gaelic', 'sr': 'serbian', 'st': 'sesotho', 'sn': 'shona', 'sd': 'sindhi',
    'si': 'sinhala', 'sk': 'slovak', 'sl': 'slovenian', 'so': 'somali', 'es': 'spanish', 'su': 'sundanese', 'sw': 'swahili',
    'sv': 'swedish', 'tg': 'tajik', 'ta': 'tamil', 'te': 'telugu', 'th': 'thai', 'tr': 'turkish', 'uk': 'ukrainian',
    'ur': 'urdu', 'ug': 'uyghur', 'uz': 'uzbek', 'vi': 'vietnamese', 'cy': 'welsh', 'xh': 'xhosa', 'yi': 'yiddish',
    'yo': 'yoruba', 'zu': 'zulu',
}

def translate(request):
    if request.method == 'POST':
        text = request.POST.get('text')
        language = request.POST.get('target_language')  # Match 'name' in the HTML form

        if not text:
            response_data = {"error": "No text provided"}
            return HttpResponse(json.dumps(response_data), content_type="application/json", status=400)

        translator = Translator()
        try:
            translated = translator.translate(text, dest=language)

            # Pass the result to the template
            return render(request, 'translate.html', {
                'original_text': text,
                'translated_text': translated.text,
                'selected_language': language,
                'languages': LANGUAGES,
            })
        except Exception as e:
            return render(request, 'translate.html', {
                'error': f"Translation failed: {e}",
                'languages': LANGUAGES,
            })

    # Render the page with a blank form on GET request
    return render(request, 'translate.html', {
        'languages': LANGUAGES
    })


@login_required
def post_doubt(request):
    if request.method == 'POST':
        # Get the form data from the POST request
        doubt_text = request.POST.get('doubt')
        teacher_id = request.POST.get('teacher')  
        teacher = User.objects.get(id=teacher_id)
        Doubts.objects.create(doubt=doubt_text, posted_by=request.user, posted_to=teacher)
        return redirect('doubts_posted')  # Redirect to a page showing posted doubts

    # Retrieve the list of teachers (Profiles with role "teacher")
    teachers = Profile.objects.filter(role="teacher")
    return render(request, 'post_doubt.html', {'teachers': teachers})


@login_required
def doubts_posted(request):
    doubts_posted = Doubts.objects.filter(posted_by=request.user)
    return render(request, 'doubts_posted.html', {'doubts_posted': doubts_posted})

@login_required
def doubts_recieved(request):
    doubts_recieved = Doubts.objects.filter(posted_to=request.user)
    return render(request, 'doubts_recieved.html', {'doubts_recieved': doubts_recieved})

def purpose(request):
    return render(request, 'purpose.html')

def board(request):
    # Call virtual_painter() as a separate process
    subprocess.Popen(["python", "signup/VirtualPainter.py"])
    return redirect("purpose")


def upload_to_drive(file):
    """
    Upload the file to Google Drive and return the file link.
    """
    # Define file metadata
    file_metadata = {'name': file.name}
    
    # Create a media upload object
    media = MediaIoBaseUpload(file, mimetype='application/octet-stream', resumable=True)
    
    # Upload the file
    uploaded_file = service.files().create(
        body=file_metadata,
        media_body=media,
        fields='id'
    ).execute()

    # Get file ID
    file_id = uploaded_file.get('id')

    # Set permissions to make the file accessible by anyone
    permission = {
        'type': 'anyone',  # Make the file accessible to anyone
        'role': 'reader'   # Read-only access
    }
    service.permissions().create(
        fileId=file_id,
        body=permission
    ).execute()

    # Generate the access link
    file_link = f"https://drive.google.com/file/d/{file_id}/view"
    
    return file_link

def delete_from_drive(file_link):
    """
    Deletes a file from Google Drive using its link.
    """
    try:
        file_id = file_link.split('/d/')[1].split('/')[0]
        service.files().delete(fileId=file_id).execute()
    except Exception as e:
        print(f"Error deleting from Google Drive: {e}")
        raise

@login_required
def upload(request):
    if request.method == 'POST':
        file_name = request.POST['file_name']
        uploaded_file = request.FILES['file']

        if not uploaded_file:
            messages.error(request, "No file uploaded.")
            return redirect('upload_notes_or_class')

        # Define valid extensions for Notes and Class categories
        notes_extensions = ['.pdf', '.docx', '.pptx', '.txt']
        class_extensions = ['.mp4', '.avi', '.mkv', '.mov', '.flv', '.wmv']
        file_extension = os.path.splitext(uploaded_file.name)[1].lower()

        # Determine the category and upload to Google Drive
        if file_extension in notes_extensions:
            file_link = upload_to_drive(uploaded_file)
            Notes.objects.create(
                note_name=file_name,
                note_link=file_link,
                created_by=request.user
            )
            messages.success(request, "Note uploaded successfully.")
        
        elif file_extension in class_extensions:
            file_link = upload_to_drive(uploaded_file)
            Classes.objects.create(
                class_name=file_name,
                class_link=file_link,
                created_by=request.user
            )
            messages.success(request, "Class video uploaded successfully.")
        
        else:
            messages.error(request, "Invalid file type. Please upload a valid Notes or Class file.")
            return redirect('upload')

        return redirect('uploads')
    return render(request, 'upload.html')

@login_required
def uploads(request):
    user = request.user
    selected_option = request.GET.get('option', 'classes')  # Default to 'classes' if no option selected

    # Fetch data based on the selected option
    if selected_option == 'notes':
        data = Notes.objects.filter(created_by=user)
    else:
        data = Classes.objects.filter(created_by=user)

    # Handle form submission for deletion
    if request.method == 'POST' and 'delete_item' in request.POST:
        item_id = request.POST.get('item_id')
        item_type = request.POST.get('item_type')  # New field to identify type
        try:
            if not item_id or not item_type:
                raise ValueError("Item ID and type are required.")
            
            if item_type == 'notes':
                item = Notes.objects.get(id=item_id, created_by=user)
                delete_from_drive(item.note_link)  # Delete from Google Drive
                item.delete()  # Delete from the database
            elif item_type == 'classes':
                item = Classes.objects.get(id=item_id, created_by=user)
                delete_from_drive(item.class_link)  # Delete from Google Drive
                item.delete()  # Delete from the database
            else:
                raise ValueError("Invalid item type.")

            messages.success(request, "Item deleted successfully.")
            return redirect('uploads')
        except Exception as e:
            messages.error(request, f"An error occurred during deletion: {e}")
            return redirect('uploads')

    context = {
        'selected_option': selected_option,
        'data': data,
    }
    return render(request, 'uploads.html', context)

@login_required
def notes_classes(request):
    selected_option = request.GET.get('option', 'classes')  # Default to 'classes'
    if selected_option == 'notes':
        data = Notes.objects.all()
    else:
        data = Classes.objects.all()

    return render(request, 'notes_classes.html', {
        'selected_option': selected_option,
        'data': data
    })


@login_required
def teacher_home(request):
    return render(request, 'teacher_home.html')

@login_required
def student_home(request):
    return render(request, 'student_home.html')

@login_required
def create_test(request):
    if request.method == 'POST':
        try:
            test_name = request.POST['test_name']
            file = request.FILES['file']
            test_open_time = make_aware(datetime.fromisoformat(request.POST['test_open_time']))
            test_close_time = make_aware(datetime.fromisoformat(request.POST['test_close_time']))
            test_time_span = int(request.POST['test_time_span'])
            num_questions = int(request.POST['num_questions'])

            if test_open_time >= test_close_time:
                return render(request, 'teacher_home.html', {'error': 'Test open time must be before close time.'})
            if num_questions < 1:
                return render(request, 'teacher_home.html', {'error': 'Number of questions must be at least 1.'})

            with transaction.atomic():
                test = Test.objects.create(
                    test_name=test_name,
                    open_time=test_open_time,
                    close_time=test_close_time,
                    duration=test_time_span,
                    num_questions=num_questions,
                    created_by=request.user
                )
                text = extract_text_from_file(file)
                if not text:
                    return render(request, 'teacher_home.html', {'error': 'Unsupported file format or empty file.'})

                mcqs = Question_mcqs_generator(text, num_questions)
                mcq_blocks = mcqs.strip().split("\n\n")
                for mcq_block in mcq_blocks:
                    lines = mcq_block.strip().split("\n")
                    if len(lines) >= 6:
                        question_text = lines[1].replace("Question:", "").strip()
                        options = {
                            "A": lines[2].split(") ")[1].strip(),
                            "B": lines[3].split(") ")[1].strip(),
                            "C": lines[4].split(") ")[1].strip(),
                            "D": lines[5].split(") ")[1].strip(),
                        }
                        correct_option = next(
                            (line.replace("Correct Answer:", "").strip() for line in lines if line.startswith("Correct Answer:")),
                            None
                        )
                        if correct_option and correct_option in options:
                            TestQuestion.objects.create(
                                test=test,
                                question_text=question_text,
                                option_a=options["A"],
                                option_b=options["B"],
                                option_c=options["C"],
                                option_d=options["D"],
                                correct_answer=correct_option
                            )
            return redirect("tests_created")
        except Exception as e:
            return render(request, 'teacher_home.html', {'error': f"An error occurred: {e}"})
    return render(request, 'create_test.html')

@login_required
def tests_created(request):
    tests = Test.objects.filter(created_by=request.user)
    return render(request, 'tests_created.html', {'tests': tests})

@login_required
def attempt_test(request):
    if request.method == 'POST':
        try:
            test_id = request.POST['test_id']
            test_attempts = StudentTestAttempt.objects.filter(student=request.user, test_id=test_id)
            if test_attempts.exists():
                # If the student has already attempted the test, show an error message
                return render(request, 'attempt_test.html', {
                    'error': "You have already attempted this test. Multiple attempts are not allowed."
                })
            
            # If no previous attempts, allow the student to proceed to the test
            return redirect('test', test_id=test_id)
        
        except Exception as e:
            # Handle any exceptions and show a user-friendly error message
            return render(request, 'attempt_test.html', {
                'error': f"An error occurred: {e}"
            })
    
    # Render the attempt test page for GET requests
    return render(request, 'attempt_test.html')

@login_required
def test(request, test_id):
    try:
        test = Test.objects.get(id=test_id)
        if not test.open_time <= now() <= test.close_time:
            return HttpResponse("This test is not available at the moment.", status=400)

        questions = TestQuestion.objects.filter(test=test)
        context = {
            'TestQuestions': questions,
            'test_id': test_id,
            'test_name': test.test_name,
            'test_duration_seconds': test.duration * 60,
        }
        return render(request, 'test.html', context)
    except Test.DoesNotExist:
        return HttpResponse("Test not found.", status=404)

@login_required
def submit_test(request, test_id):
    if request.method == "POST":
        try:
            test = Test.objects.get(id=test_id)
            questions = TestQuestion.objects.filter(test=test)
            student = request.user

            correct_answers = sum(
                1 for question in questions
                if request.POST.get(f"question_{question.id}") == question.correct_answer
            )
            total_questions = questions.count()
            marks_obtained = correct_answers

            StudentMarks.objects.update_or_create(
                test=test,
                student=student,
                defaults={'marks_obtained': marks_obtained}
            )
            StudentTestAttempt.objects.update_or_create(
                student=student,
                test=test,
                defaults={
                    'correct_answers': correct_answers,
                    'total_questions': total_questions,
                    'submitted_at': now()
                }
            )
            return redirect("tests_attempted")
        except Test.DoesNotExist:
            return HttpResponse("Test not found.", status=404)
    return HttpResponse("Invalid request method.", status=405)

@login_required
def tests_attempted(request):
    test_attempts = StudentTestAttempt.objects.filter(student=request.user).select_related('test')
    return render(request, 'tests_attempted.html', {'test_attempts': test_attempts})

@login_required
def download_marks(request, test_id):
    try:
        test = Test.objects.get(id=test_id, created_by=request.user)
        if test.close_time > now():
            return HttpResponse("Marks will be available after the exam ends.", status=400)

        marks = StudentMarks.objects.filter(test=test).select_related('student')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Marks for {test.test_name}"

        sheet.merge_cells('A1:C1')
        title_cell = sheet['A1']
        title_cell.value = f"Marks for Test: {test.test_name}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        headers = ["Student Username", "Marks Obtained"]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row_num, mark in enumerate(marks, start=3):
            sheet.cell(row=row_num, column=1).value = mark.student.username
            sheet.cell(row=row_num, column=2).value = mark.marks_obtained

        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 20

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="marks_test_{test_id}.xlsx"'
        workbook.save(response)
        return response
    except Test.DoesNotExist:
        return HttpResponse("Test not found.", status=404)

@login_required
def download_questions(request, test_id):
    try:
        test = Test.objects.get(id=test_id, created_by=request.user)
        questions = TestQuestion.objects.filter(test=test)
        question_data = [
            {
                "question_text": question.question_text,
                "option_a": question.option_a,
                "option_b": question.option_b,
                "option_c": question.option_c,
                "option_d": question.option_d,
                "correct_answer": question.correct_answer,
            } for question in questions
        ]
        response = HttpResponse(json.dumps(question_data, indent=4), content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="questions_test_{test_id}.json"'
        return response
    except Test.DoesNotExist:
        return HttpResponse("Test not found.", status=404)

@login_required
def preperation(request):
    if request.method == 'POST':
        try:
            # Handle file and number of questions input
            file = request.FILES['file']
            num_questions = int(request.POST['num_questions'])

            if num_questions < 1:
                return render(request, 'preperation.html', {'error': 'Number of questions must be at least 1.'})

            # Extract text from file
            text = extract_text_from_file(file)
            if not text:
                return render(request, 'preperation.html', {'error': 'Unsupported file format or empty file.'})

            # Generate MCQs
            mcqs = Question_mcqs_generator(text, num_questions)

            # Render the page with MCQs and provide PDF link
            return render(request, 'preperation.html', {'mcqs': mcqs, 'num_questions': num_questions})

        except Exception as e:
            return render(request, 'preperation.html', {'error': f"An error occurred: {e}"})
    
    return render(request, 'preperation.html')


@login_required
def generate_pdf(request, num_questions, questions_and_answers):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    c.setFont("Helvetica", 12)
    c.drawString(100, height - 50, "Generated Questions and Answers:")

    y_position = height - 70
    for qna in questions_and_answers:
        c.drawString(100, y_position, f"Q: {qna['question_text']}")
        y_position -= 20
        for option, answer in qna['options'].items():
            c.drawString(120, y_position, f"{option}. {answer}")
            y_position -= 15
        c.drawString(100, y_position, f"Correct Answer: {qna['correct_answer']}")
        y_position -= 30

        if y_position < 100:  # If space is running out, add a new page
            c.showPage()
            c.setFont("Helvetica", 12)
            y_position = height - 50

    c.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="questions_and_answers.pdf"'
    return response


def signup(request):
    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        role = request.POST.get("role")
        password = request.POST.get("password")

        if User.objects.filter(username=username).exists():
            return render(request, "signup.html", {"error": "Username already taken."})
        if User.objects.filter(email=email).exists():
            return render(request, "signup.html", {"error": "Email already registered."})

        user = User.objects.create_user(username=username, email=email, password=password)
        Profile.objects.create(user=user, role=role)

        return redirect("login")
    return render(request, "signup.html")

def login_user(request):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(username=username, password=password)
        if user is not None:
            auth_login(request, user)
            profile = Profile.objects.get(user=user)
            if profile.role == 'student':
                return redirect("student_home")
            elif profile.role == 'teacher':
                return redirect("teacher_home")
        else:
            return render(request, "login.html", {"error": "Invalid login credentials"})

    return render(request, "login.html")

def logout_user(request):
    auth_logout(request)
    return redirect("login")
